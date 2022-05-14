from openpyxl import load_workbook

excel_filename = "Exported_data.xlsx"
wb = load_workbook(filename=excel_filename)
print(wb.sheetnames)

# Get the sheet by name
worksheet = wb["All points"]

print(worksheet.title)

# get all the rows
filename = "measurements.txt"
with open(filename, 'w') as f:
    # enumerate the rows
    for i, row in enumerate(worksheet.iter_rows()):
        line = ""
        for cell in row:
            line += str(cell.value) + ":"
        line = line.strip(":")
        line += "\n"
        if i != 0:  # skip the header
            f.write(line)

wb.close()
